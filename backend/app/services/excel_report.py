from collections.abc import Mapping
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


SUMMARY_HEADERS = [
    "Student Code",
    "Student Name",
    "Class ID",
    "Total Sessions",
    "Present",
    "Late",
    "Absent",
    "Absence Rate (%)",
    "Absence Alert",
]
DETAIL_HEADERS = [
    "Session ID",
    "Course Name",
    "Start Time",
    "End Time",
    "Status",
]


def build_attendance_report(
    statistics: Mapping[str, Any],
    *,
    student_name: str | None = None,
) -> BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Attendance Summary"
    details = workbook.create_sheet("Session Details")

    summary.append(SUMMARY_HEADERS)
    summary.append(
        [
            statistics.get("student_code"),
            student_name or "",
            statistics.get("class_id"),
            statistics.get("total_sessions", 0),
            statistics.get("present", 0),
            statistics.get("late", 0),
            statistics.get("absent", 0),
            statistics.get("absence_rate", 0),
            "Yes" if statistics.get("absence_alert", False) else "No",
        ]
    )

    details.append(DETAIL_HEADERS)
    for session in statistics.get("details", []):
        details.append(
            [
                session.get("session_id"),
                session.get("course_name"),
                session.get("start_time"),
                session.get("end_time"),
                session.get("status"),
            ]
        )

    for worksheet in (summary, details):
        _format_worksheet(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _format_worksheet(worksheet: Any) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 30)
