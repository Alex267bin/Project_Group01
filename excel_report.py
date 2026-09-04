from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def create_attendance_report(data, filename="attendance_report.xlsx"):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.title = "Attendance Report"

    headers = [
        "Student ID",
        "Student Code",
        "Student Name",
        "Total Sessions",
        "Present",
        "Absent",
        "Late",
        "Attendance Rate (%)",
        "Absence Rate (%)"
    ]

    # Tạo tiêu đề
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=1,
            column=column,
            value=header
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Thêm dữ liệu
    for row_number, student in enumerate(data, start=2):
        worksheet.cell(row=row_number, column=1, value=student["student_id"])
        worksheet.cell(row=row_number, column=2, value=student["student_code"])
        worksheet.cell(row=row_number, column=3, value=student["student_name"])
        worksheet.cell(row=row_number, column=4, value=student["total_sessions"])
        worksheet.cell(row=row_number, column=5, value=student["present"])
        worksheet.cell(row=row_number, column=6, value=student["absent"])
        worksheet.cell(row=row_number, column=7, value=student["late"])
        worksheet.cell(row=row_number, column=8, value=student["attendance_rate"])
        worksheet.cell(row=row_number, column=9, value=student["absence_rate"])

    # Tự điều chỉnh độ rộng cột
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save(filename)

    return filename